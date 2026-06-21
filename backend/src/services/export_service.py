"""Export Service for various formats"""
from typing import List, Dict, Optional
from io import StringIO, BytesIO
from datetime import datetime
import csv


class ExportService:
    """Service for exporting screening results"""
    
    @staticmethod
    def export_to_csv(results: List[Dict]) -> str:
        """Export results to CSV format
        
        Args:
            results: List of screening result dicts
        
        Returns:
            CSV string
        """
        if not results:
            return ""
        
        output = StringIO()
        fieldnames = [
            'Rank', 'Symbol', 'Company', 'Price', 'Change %',
            'Volume', 'Market Cap', 'AI Score', 'Recommendation',
            'Support', 'Resistance', 'Stop Loss', 'Target1', 'Target2',
            'Risk/Reward', 'EMA Alignment', 'RSI', 'MACD', 'ADX',
            'Relative Volume', 'Trading Style'
        ]
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for result in results:
            writer.writerow({
                'Rank': result.get('rank', ''),
                'Symbol': result.get('symbol', ''),
                'Company': result.get('company_name', ''),
                'Price': result.get('close_price', ''),
                'Change %': result.get('change_percent', ''),
                'Volume': result.get('volume', ''),
                'Market Cap': result.get('market_cap', ''),
                'AI Score': result.get('ai_score', ''),
                'Recommendation': result.get('recommendation', ''),
                'Support': result.get('support', ''),
                'Resistance': result.get('resistance', ''),
                'Stop Loss': result.get('stop_loss', ''),
                'Target1': result.get('target_1', ''),
                'Target2': result.get('target_2', ''),
                'Risk/Reward': result.get('risk_reward_ratio', ''),
                'EMA Alignment': result.get('ema_alignment_score', ''),
                'RSI': result.get('rsi_score', ''),
                'MACD': result.get('macd_score', ''),
                'ADX': result.get('adx_score', ''),
                'Relative Volume': result.get('relative_volume_score', ''),
                'Trading Style': result.get('trading_style', '')
            })
        
        return output.getvalue()
    
    @staticmethod
    def export_to_excel(results: List[Dict]) -> Optional[bytes]:
        """Export results to Excel format
        
        Args:
            results: List of screening result dicts
        
        Returns:
            Excel file bytes or None
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Screening Results"
            
            # Define headers
            headers = [
                'Rank', 'Symbol', 'Company', 'Price', 'Change %',
                'Volume', 'Market Cap', 'AI Score', 'Recommendation',
                'Support', 'Resistance', 'Stop Loss', 'Target1', 'Target2',
                'Risk/Reward', 'EMA Alignment', 'RSI', 'MACD', 'ADX',
                'Relative Volume', 'Trading Style'
            ]
            
            # Add headers
            for col, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add data rows
            for row_idx, result in enumerate(results, start=2):
                values = [
                    result.get('rank', ''),
                    result.get('symbol', ''),
                    result.get('company_name', ''),
                    result.get('close_price', ''),
                    result.get('change_percent', ''),
                    result.get('volume', ''),
                    result.get('market_cap', ''),
                    result.get('ai_score', ''),
                    result.get('recommendation', ''),
                    result.get('support', ''),
                    result.get('resistance', ''),
                    result.get('stop_loss', ''),
                    result.get('target_1', ''),
                    result.get('target_2', ''),
                    result.get('risk_reward_ratio', ''),
                    result.get('ema_alignment_score', ''),
                    result.get('rsi_score', ''),
                    result.get('macd_score', ''),
                    result.get('adx_score', ''),
                    result.get('relative_volume_score', ''),
                    result.get('trading_style', '')
                ]
                
                for col_idx, value in enumerate(values, start=1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Adjust column widths
            for col in range(1, len(headers) + 1):
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
            
            # Save to bytes
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.getvalue()
        
        except ImportError:
            return None
    
    @staticmethod
    def export_to_pdf(results: List[Dict], filename: str = 'screening_results.pdf') -> Optional[bytes]:
        """Export results to PDF format
        
        Args:
            results: List of screening result dicts
            filename: Output filename
        
        Returns:
            PDF file bytes or None
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=landscape(letter))
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1F4E78'),
                alignment=1,
                spaceAfter=20
            )
            title = Paragraph('Stock Screener Analysis Report', title_style)
            elements.append(title)
            
            # Date
            date_text = Paragraph(
                f'Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")}',
                styles['Normal']
            )
            elements.append(date_text)
            elements.append(Spacer(1, 0.3*inch))
            
            # Results table
            if results:
                headers = [
                    'Rank', 'Symbol', 'Company', 'Price', 'Score', 'Rec.',
                    'Support', 'Resistance', 'Stop Loss', 'Target1', 'Target2'
                ]
                
                data = [headers]
                for result in results:
                    data.append([
                        str(result.get('rank', '')),
                        str(result.get('symbol', '')),
                        str(result.get('company_name', ''))[:20],
                        f"{result.get('close_price', '')}B",
                        str(result.get('ai_score', '')),
                        str(result.get('recommendation', ''))[:4],
                        str(result.get('support', '')),
                        str(result.get('resistance', '')),
                        str(result.get('stop_loss', '')),
                        str(result.get('target_1', '')),
                        str(result.get('target_2', ''))
                    ])
                
                table = Table(data, colWidths=[0.4*inch]*len(headers))
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                ]))
                elements.append(table)
            
            doc.build(elements)
            output.seek(0)
            return output.getvalue()
        
        except ImportError:
            return None
